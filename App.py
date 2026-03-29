from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os
from functools import wraps

app = Flask(__name__)
app.secret_key = "chave_seguranca_inspecao_pro"

ARQUIVO = "base_dados.xlsx"
UPLOAD = "uploads"

if not os.path.exists(UPLOAD):
    os.makedirs(UPLOAD)

# -------------------------
# DECORADORES DE ACESSO
# -------------------------

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

# -------------------------
# FUNÇÕES DE APOIO
# -------------------------

def verificar_arquivo_excel():
    if not os.path.exists(ARQUIVO):
        wb = Workbook()
        ws_eq = wb.active
        ws_eq.title = "Equipamentos"
        ws_eq.append(["Equipamento", "Item", "Tipo", "Min", "Max", "Setor", "Prioridade"])
        
# Localize este bloco dentro de verificar_arquivo_excel e substitua:
        wb.create_sheet("Inspecoes").append([
            "ID", "Data", "Equipamento", "Item", "Valor", "Setor", 
            "Status_Inspecao", "Relato_Inspecao", "Foto", "Usuario", 
            "Causa_Raiz", "Acao_Tomada", "Status_Tratativa", "Latitude", "Longitude", 
            "OS", "Status_OS", "Nivel_alarme" 
        ])
        
        wb.create_sheet("Programacao").append(["Data", "Equipamento", "Turno", "Setor"])
        
        ws_user = wb.create_sheet("Usuarios")
        ws_user.append(["Usuario", "Senha", "Nível", "Setor"])
        ws_user.append(["admin", "123", "ADMIN", "PRODUCAO"])
        
        wb.save(ARQUIVO)

def obter_turno_atual():
    agora = datetime.now().time()
    t1_ini = datetime.strptime("08:00", "%H:%M").time()
    t1_fim = datetime.strptime("15:45", "%H:%M").time()
    t2_ini = datetime.strptime("16:00", "%H:%M").time()
    t2_fim = datetime.strptime("23:45", "%H:%M").time()
    t3_ini = datetime.strptime("00:00", "%H:%M").time()
    t3_fim = datetime.strptime("07:45", "%H:%M").time()

    if t3_ini <= agora <= t3_fim: return 3
    if t1_ini <= agora <= t1_fim: return 1
    if t2_ini <= agora <= t2_fim: return 2
    if agora < t1_ini: return 1
    if agora < t2_ini: return 2
    return 3

# -------------------------
# ROTAS
# -------------------------

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD, filename)

@app.route("/login", methods=["GET", "POST"])
def login():
    if "usuario" in session: return redirect(url_for("index"))
    if request.method == "POST":
        u, s = request.form.get("usuario"), request.form.get("senha")
        wb = load_workbook(ARQUIVO, data_only=True)
        aba = wb["Usuarios"]
        for l in aba.iter_rows(min_row=2, values_only=True):
            if l and str(l[0]) == u and str(l[1]) == s:
                session.update({
                    "usuario": l[0], 
                    "nivel": str(l[2]).upper().strip(), 
                    "setor": str(l[3]).upper().strip()
                })
                return redirect(url_for("index"))
        flash("Usuário ou senha inválidos.")
    return render_template("login.html")

# ===================
# LOGIN REQUISIÇÃO
# ===================

@app.route("/")
@login_required
def index():
    wb = load_workbook(ARQUIVO, data_only=True)
    setor_usuario = str(session.get("setor", "")).strip().upper()
    nivel_usuario = session.get("nivel")
    
    # Regra de Acesso Full (Visualização)
    acesso_total = (nivel_usuario in ["ADMIN", "ADMINFULL"]) or (setor_usuario == "GERAL")
    # Regra de Bloqueio de Preenchimento para ADMINFULL
    bloqueio_preenchimento = (nivel_usuario == "ADMINFULL")

    data_hoje_obj = datetime.now().date()
    data_hoje_str = data_hoje_obj.strftime("%d/%m/%Y")
    turno_atual = obter_turno_atual()

    total_48h = 0
    pendentes_48h = 0
    alertas_48h = 0
    tratativas_48h = 0
    inspecionados_hoje = []
    
    # --- NOVA VARIÁVEL PARA ALARMES ATIVOS ---
    alarmes_ativos = {} 

if "Inspecoes" in wb.sheetnames:
        for l in wb["Inspecoes"].iter_rows(min_row=2, values_only=True):
            # --- INÍCIO DO BLOCO SEGURO ---
            try:
                # Se a linha estiver vazia ou a data for nula, pula para a próxima
                if not l or l[1] is None: 
                    continue 

                if isinstance(l[1], datetime):
                    dt_insp_obj = l[1].date()
                else:
                    # Tenta converter o texto em data, se falhar, o 'except' captura
                    data_texto = str(l[1]).split()[0]
                    dt_insp_obj = datetime.strptime(data_texto, "%d/%m/%Y").date()
            except Exception:
                # Se der qualquer erro na linha (data errada, formato doido),
                # ele ignora essa linha e continua o resto do site.
                continue 
           
            dias_diff = (data_hoje_obj - dt_insp_obj).days
            if 0 <= dias_diff <= 1:
                setor_insp = str(l[5]).strip().upper()
                if acesso_total or setor_insp == setor_usuario:
                    total_48h += 1
                    if str(l[6]).upper() == "ALARME":
                        alertas_48h += 1
                        if str(l[12]).upper() == "PENDENTE":
                            tratativas_48h += 1
                    if dt_insp_obj == data_hoje_obj:
                        inspecionados_hoje.append(str(l[2]))
            
            # --- AJUSTE: CAPTURA DE ALARMES ATIVOS COM OS ---
            status_inspecao = str(l[6]).upper() if l[6] else ""
            status_os = str(l[16]).upper() if len(l) > 16 and l[16] else "PENDENTE"
            
            if status_inspecao == "ALARME" and status_os != "CONCLUÍDO":
                eq_nome_insp = str(l[2])
                os_num = l[15] if l[15] else "S/ OS"
                alarmes_ativos[eq_nome_insp] = os_num

    if "Programacao" in wb.sheetnames:
        for l in wb["Programacao"].iter_rows(min_row=2, values_only=True):
            if not l or not l[0]: continue
            try:
                dt_prog_obj = l[0].date() if isinstance(l[0], datetime) else datetime.strptime(str(l[0]), "%d/%m/%Y").date()
            except: continue

            dias_prog = (data_hoje_obj - dt_prog_obj).days
            if 0 <= dias_prog <= 1:
                setor_p = str(l[3]).strip().upper()
                if acesso_total or setor_p == setor_usuario:
                    if str(l[1]) not in inspecionados_hoje:
                        pendentes_48h += 1

    equipamentos_permitidos = []
    if "Programacao" in wb.sheetnames:
        for l in wb["Programacao"].iter_rows(min_row=2, values_only=True):
            if not l or not l[0]: continue
            dt_planilha = l[0].strftime("%d/%m/%Y") if isinstance(l[0], datetime) else str(l[0])
            t_raw = str(l[2]).upper().replace('T', '').strip() if l[2] else "0"
            try: t_planilha = int(t_raw)
            except: t_planilha = 0
            setor_planilha = str(l[3]).strip().upper()

            if dt_planilha == data_hoje_str and t_planilha == turno_atual:
                eq_nome = str(l[1])
                if (acesso_total or setor_planilha == setor_usuario) and (eq_nome not in inspecionados_hoje):
                    if eq_nome not in equipamentos_permitidos:
                        equipamentos_permitidos.append(eq_nome)

    equipamentos_filtrados = {}
    if "Equipamentos" in wb.sheetnames:
        for linha in wb["Equipamentos"].iter_rows(min_row=2, values_only=True):
            if linha and str(linha[0]) in equipamentos_permitidos:
                eq = str(linha[0]).strip()
                if eq not in equipamentos_filtrados: equipamentos_filtrados[eq] = []
                equipamentos_filtrados[eq].append({
                    "item": linha[1], "tipo": str(linha[2]).upper(), 
                    "min": linha[3], "max": linha[4],
                    "prioridade": linha[6] if len(linha) > 6 else "Normal"
                })
    
    selecionado = request.args.get("equipamento")
    itens = equipamentos_filtrados.get(selecionado, [])
    
    return render_template("index.html", 
                           equipamentos=equipamentos_filtrados, 
                           alarmes_ativos=alarmes_ativos, # Nova variável integrada
                           itens=itens, 
                           selecionado=selecionado, 
                           turno=turno_atual,
                           total_periodo=total_48h,
                           pendentes_periodo=pendentes_48h,
                           alertas_periodo=alertas_48h,
                           tratativas_periodo=tratativas_48h,
                           bloqueio=bloqueio_preenchimento)

# ===================
# MAPA
# ===================

@app.route("/mapa")
@login_required
def mapa():
    wb = load_workbook(ARQUIVO, data_only=True)
    setor_usuario = str(session.get("setor", "")).strip().upper()
    nivel_usuario = session.get("nivel")
    acesso_total = (nivel_usuario in ["ADMIN", "ADMINFULL"]) or (setor_usuario == "GERAL")
    
    pontos_mapa = {}
    if "Inspecoes" in wb.sheetnames:
        aba = wb["Inspecoes"]
        for row in aba.iter_rows(min_row=2, values_only=True):
            if not row or not row[13] or not row[14]: continue
            setor_insp = str(row[5]).upper().strip()
            if acesso_total or setor_insp == setor_usuario:
                eq = str(row[2])
                status_atual = str(row[6]).upper()
                relato = row[7] if row[7] else "Sem relato"
                valor = row[4] if row[4] else "N/A"
                if eq not in pontos_mapa:
                    pontos_mapa[eq] = {"lat": row[13], "lon": row[14], "eq": eq, "setor": setor_insp, "status": status_atual, "item": row[3], "valor": valor, "Relato_Inspecao": relato, "data": str(row[1])}
                if status_atual == "ALARME":
                    pontos_mapa[eq].update({"status": "ALARME", "Relato_Inspecao": relato, "valor": valor})

    return render_template("mapa.html", pontos=list(pontos_mapa.values()))

# ================
# PROGRAMAÇÃO
# ================

@app.route("/programacao")
@login_required
def programacao():
    wb = load_workbook(ARQUIVO, data_only=True)
    setor_usuario = str(session.get("setor", "")).strip().upper()
    nivel_usuario = session.get("nivel")
    acesso_total = (nivel_usuario in ["ADMIN", "ADMINFULL"]) or (setor_usuario == "GERAL")
    bloqueio_preenchimento = (nivel_usuario == "ADMINFULL")
    turno_atual = obter_turno_atual()
    usuario = {"nome": session.get("usuario"), "setor": setor_usuario, "nivel": nivel_usuario, "turno_momento": turno_atual}
    
    realizados = {}
    if "Inspecoes" in wb.sheetnames:
        for l in wb["Inspecoes"].iter_rows(min_row=2, values_only=True):
            if l and l[1] and l[2]:
                dt_insp = str(l[1]).split()[0]
                eq_insp = str(l[2])
                if dt_insp not in realizados: realizados[dt_insp] = []
                realizados[dt_insp].append(eq_insp)

# ===========
# AGENDA
# ===========

    agenda = []
    if "Programacao" in wb.sheetnames:
        for l in wb["Programacao"].iter_rows(min_row=2, values_only=True):
            if l and l[0]:
                dt_val = l[0].strftime("%d/%m/%Y") if isinstance(l[0], datetime) else str(l[0])
                eq_nome = str(l[1])
                setor_p = str(l[3]).upper().strip() if l[3] else ""
                if acesso_total or setor_p == setor_usuario:
                    if eq_nome not in realizados.get(dt_val, []):
                        agenda.append({"data": dt_val, "equipamento": eq_nome, "turno": l[2], "setor": setor_p})
                        
    return render_template("programacao.html", agenda=agenda, usuario=usuario, bloqueio=bloqueio_preenchimento)

# ===============
# ROTINA PCMI
# ===============

@app.route("/pcm")
@login_required
def pcm():
    if session.get("nivel") != "ADMINFULL":
        flash("Acesso restrito.")
        return redirect(url_for("index"))

    from openpyxl import load_workbook
    import os

    # Carrega o workbook (usamos o caminho padrão ARQUIVO)
    wb = load_workbook(ARQUIVO)
    aba = wb["Inspecoes"]

    alertas_dict = {}
    memoria_os = {}
    status_os_map = {}

    # 🔥 1. MAPEAR STATUS DAS OS (COLUNA P=15, Q=16)
    # Primeiro passamos para entender quais OS já foram concluídas no sistema
    for row in aba.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 17: continue
        os_cell = str(row[15]).strip() if row[15] else None
        status_cell = str(row[16]).strip().upper() if row[16] else "PENDENTE"

        if os_cell and os_cell != "NONE":
            # Se já houver um "CONCLUÍDO" para esta OS em qualquer linha, prevalece
            if status_os_map.get(os_cell) != "CONCLUÍDO":
                status_os_map[os_cell] = status_cell

    # 🔥 2. ATUALIZAÇÃO DE STATUS NO EXCEL (SE NECESSÁRIO) E COLETA DE DADOS
    # Vamos percorrer as linhas para preencher a memória e o dicionário de exibição
    # Usamos o iter_rows normal para poder editar as células se necessário
    for row_idx, row in enumerate(aba.iter_rows(min_row=2), start=2):
        # Convertemos para lista de valores para facilitar a leitura rápida
        valores = [cell.value for cell in row]
        if not valores or len(valores) < 7: continue

        # Identificação básica
        equipamento = str(valores[2]).strip() if valores[2] else ""
        item = str(valores[3]).strip() if valores[3] else ""
        chave = f"{equipamento}_{item}" # Chave que causava o KeyError
        
        status_inspecao = str(valores[6]).upper().strip() if valores[6] else ""
        os_gravada = str(valores[15]).strip() if valores[15] else ""
        
        # Sincroniza o status da OS na planilha se o mapa indicar CONCLUÍDO
        if os_gravada in status_os_map:
            status_atualizado = status_os_map[os_gravada]
            if valores[16] != status_atualizado:
                aba.cell(row=row_idx, column=17).value = status_atualizado
                status_os = status_atualizado
            else:
                status_os = str(valores[16]).upper().strip() if valores[16] else "PENDENTE"
        else:
            status_os = str(valores[16]).upper().strip() if valores[16] else "PENDENTE"

        # Captura o nível de alarme (Coluna R = índice 17)
        nivel_gravado = str(valores[17]).strip().upper() if len(valores) > 17 and valores[17] else ""

        # 🧠 Alimentar memória de OS (para auto-preencher novas inspeções do mesmo item)
        if os_gravada and os_gravada != "NONE" and status_os != "CONCLUÍDO":
            memoria_os[chave] = {
                "os": os_gravada, 
                "status": status_os, 
                "nivel": nivel_gravado
            }

        # 🚩 Selecionar apenas o que deve ser exibido na tela do PCM
        # Regra: Ser um ALARME e a OS não estar CONCLUÍDA
        if status_inspecao == "ALARME" and status_os != "CONCLUÍDO":
            alertas_dict[chave] = {
                "id": valores[0],
                "data": valores[1],
                "equipamento": equipamento,
                "item": item,
                "setor": valores[5],
                "os": os_gravada if os_gravada != "NONE" else "",
                "status_os": status_os,
                "nivel_alarme": nivel_gravado
            }

    # Salva as alterações de status que fizemos no Passo 2
    wb.save(ARQUIVO)

    # 🔥 3. PROCESSAMENTO FINAL (AUTO-PREENCHIMENTO COM SEGURANÇA)
    alertas_finais = []

    for chave, dados in alertas_dict.items():
        # Se a linha atual não tem OS, mas o item existe na memória (inspeção anterior)
        # Usamos o .get() para evitar KeyError se a chave sumir por algum motivo
        info_memoria = memoria_os.get(chave)
        
        if info_memoria:
            if not dados["os"] or dados["os"] == "":
                dados["os"] = info_memoria.get("os", "")
                dados["status_os"] = info_memoria.get("status", "PENDENTE")
            
            if not dados["nivel_alarme"] or dados["nivel_alarme"] == "":
                dados["nivel_alarme"] = info_memoria.get("nivel", "")

        alertas_finais.append(dados)

    return render_template("pcm.html", alertas=alertas_finais)

# =============
# SALVAR OS
# =============

@app.route("/salvar_os", methods=["POST"])
@login_required
def salvar_os():
    id_insp = request.form.get("id_inspeção")
    num_os = request.form.get("os")
    status_os = request.form.get("status_os")
    nivel_alarme = request.form.get("nivel_alarme") 

    wb = load_workbook(ARQUIVO)
    aba = wb["Inspecoes"]
    for row in range(2, aba.max_row + 1):
        if str(aba.cell(row=row, column=1).value) == str(id_insp):
            aba.cell(row=row, column=16).value = num_os # Coluna P
            aba.cell(row=row, column=17).value = status_os # Coluna Q
            aba.cell(row=row, column=18).value = nivel_alarme # Coluna R
            break
    wb.save(ARQUIVO)
    flash("Dados de OS atualizados pelo PCM!")
    return redirect(url_for("pcm"))

# ======================
# ANALISES TENDENCIA
# ======================

@app.route("/analise")
@login_required
def analise():
    wb = load_workbook(ARQUIVO, data_only=True)
    setor_usuario = str(session.get("setor", "")).strip().upper()
    nivel_usuario = session.get("nivel")
    
    acesso_total = (nivel_usuario in ["ADMIN", "ADMINFULL"]) or (setor_usuario == "GERAL")
    
    # --- 1. BUSCA DE EQUIPAMENTOS E LIMITES ---
    lista_busca = [] 
    limites_map = {} 

    if "Equipamentos" in wb.sheetnames:
        for row in wb["Equipamentos"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            
            eq_nome = str(row[0]).strip().upper()
            item_nome = str(row[2]).strip().upper()
            
            if acesso_total or str(row[1]).strip().upper() == setor_usuario:
                if str(row[0]).strip() not in lista_busca:
                    lista_busca.append(str(row[0]).strip())
                
                chave_limite = f"{eq_nome} - {item_nome}"
                try:
                    limites_map[chave_limite] = {
                        "min": float(str(row[3]).replace(',', '.')) if row[3] is not None else None,
                        "max": float(str(row[4]).replace(',', '.')) if row[4] is not None else None
                    }
                except:
                    limites_map[chave_limite] = {"min": None, "max": None}

    # --- 2. AJUSTE AQUI: CAPTURA DOS FILTROS (BOTÃO GESTÃO VS FILTRO MANUAL) ---
    # Captura via lista (vários selecionados)
    eq_selecionados = request.args.getlist("eq_busca")
    item_selecionados = request.args.getlist("item_busca")

    # Se a lista estiver vazia, tenta capturar o parâmetro individual vindo do botão da Gestão à Vista
    if not eq_selecionados and request.args.get("equipamento"):
        eq_selecionados = [request.args.get("equipamento")]
        
    if not item_selecionados and request.args.get("item"):
        item_selecionados = [request.args.get("item")]

    # --- 3. LISTA DINÂMICA DE ITENS ---
    lista_itens_disponiveis = []
    if eq_selecionados and "Inspecoes" in wb.sheetnames:
        eq_filtros_upper = [str(e).strip().upper() for e in eq_selecionados]
        for row in wb["Inspecoes"].iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5 or row[2] is None or row[4] is None: continue
            
            if str(row[2]).strip().upper() in eq_filtros_upper:
                try:
                    val_check = str(row[4]).replace(',', '.').strip()
                    float(val_check) 
                    
                    item_nome_real = str(row[3]).strip()
                    if item_nome_real and item_nome_real not in lista_itens_disponiveis:
                        lista_itens_disponiveis.append(item_nome_real)
                except (ValueError, TypeError):
                    continue

    # --- 4. COLETA DE DADOS HISTÓRICOS ---
    dados_grafico = []
    if eq_selecionados and "Inspecoes" in wb.sheetnames:
        eq_filtros_upper = [str(e).strip().upper() for e in eq_selecionados]
        item_filtros_upper = [str(i).strip().upper() for i in item_selecionados]

        for l in wb["Inspecoes"].iter_rows(min_row=2, values_only=True):
            if not l or len(l) < 5 or l[2] is None or l[3] is None: continue
            
            eq_da_linha = str(l[2]).strip().upper()
            item_da_linha = str(l[3]).strip().upper()
            
            if eq_da_linha in eq_filtros_upper:
                if not item_selecionados or item_da_linha in item_filtros_upper:
                    try:
                        val_raw = str(l[4]).replace(',', '.').strip()
                        valor_medido = float(val_raw)

                        chave_lookup = f"{eq_da_linha} - {item_da_linha}"
                        limite = limites_map.get(chave_lookup, {"min": None, "max": None})
                        
                        dados_grafico.append({
                            "data": l[1].strftime("%d/%m") if hasattr(l[1], 'strftime') else str(l[1])[:5],
                            "equipamento": str(l[2]).strip(),
                            "item": str(l[3]).strip(),
                            "valor": valor_medido,
                            "min": limite["min"],
                            "max": limite["max"]
                        })
                    except:
                        continue

    # --- 5. RETORNO PARA O TEMPLATE ---
    return render_template("analise.html", 
                           lista_busca=sorted(lista_busca), 
                           lista_itens=sorted(lista_itens_disponiveis), 
                           dados=dados_grafico, 
                           selecionados_eq=eq_selecionados,
                           selecionados_item=item_selecionados)

# ===================
# GESTÃO A VISTA
# ===================

# ===================
# GESTÃO A VISTA (CORRIGIDA)
# ===================

@app.route("/gestao_a_vista")
@login_required
def gestao_a_vista():
    from openpyxl import load_workbook
    from datetime import datetime

    wb = load_workbook(ARQUIVO, data_only=True)
    hoje = datetime.now()

    alarmes_dict = {}

    if "Inspecoes" in wb.sheetnames:
        for row in wb["Inspecoes"].iter_rows(min_row=2, values_only=True):

            if not row or len(row) < 7:
                continue

            # Status OS (Coluna Q = 16) e Status Inspeção (Coluna G = 6)
            status_os = str(row[16]).strip().upper() if len(row) > 16 and row[16] else ""
            status_inspecao = str(row[6]).strip().upper() if row[6] else ""

            # Filtra apenas o que está PENDENTE e é ALARME
            if status_os != "PENDENTE":
                continue

            if "ALARME" not in status_inspecao:
                continue

            equipamento = str(row[2]).strip().upper()
            item = str(row[3]).strip().upper()
            chave = f"{equipamento}_{item}"

            # 🔥 CAPTURA O NÍVEL DE ALARME (COLUNA R = ÍNDICE 17)
            nivel = str(row[17]).strip().upper() if len(row) > 17 and row[17] else "N/D"

            data_raw = row[1]
            data_reg = None

            if isinstance(data_raw, datetime):
                data_reg = data_raw
            else:
                try:
                    data_reg = datetime.strptime(str(data_raw), "%d/%m/%Y %H:%M")
                except:
                    try:
                        data_reg = datetime.strptime(str(data_raw), "%d/%m/%Y")
                    except:
                        continue

            valor = row[4]
            setor = row[5] if row[5] else "GERAL"

            if chave not in alarmes_dict:
                alarmes_dict[chave] = {
                    "equipamento": equipamento,
                    "item": item,
                    "valor": valor,
                    "setor": setor,
                    "data_primeira": data_reg,
                    "data_ultima": data_reg,
                    "notificacoes": 1,
                    "nivel_alarme": nivel # <--- Adicionado aqui
                }
            else:
                alarmes_dict[chave]["notificacoes"] += 1

                if data_reg < alarmes_dict[chave]["data_primeira"]:
                    alarmes_dict[chave]["data_primeira"] = data_reg

                if data_reg > alarmes_dict[chave]["data_ultima"]:
                    alarmes_dict[chave]["data_ultima"] = data_reg
                    alarmes_dict[chave]["valor"] = valor
                    # Atualiza o nível para o registro mais recente
                    alarmes_dict[chave]["nivel_alarme"] = nivel 

    lista_final = []

    for d in alarmes_dict.values():
        dias_aberto = (hoje - d["data_primeira"]).days

        lista_final.append({
            "equipamento": d["equipamento"],
            "item": d["item"],
            "valor": d["valor"],
            "setor": d["setor"],
            "data": d["data_ultima"].strftime("%d/%m/%Y %H:%M"),
            "dias": max(0, dias_aberto),
            "notificacoes": d["notificacoes"],
            "nivel_alarme": d["nivel_alarme"] # <--- Passando para a lista
        })

    lista_final = sorted(lista_final, key=lambda x: x["dias"], reverse=True)

    return render_template("gestao.html", alarmes=lista_final)

# ==================
# TRATATIVAS
# ==================

@app.route("/tratativas")
@login_required
def tratativas():
    nivel_usuario = str(session.get("nivel", "")).upper()
    setor_usuario = str(session.get("setor", "")).strip().upper()
    acesso_total = (nivel_usuario in ["ADMIN", "ADMINFULL"]) or (setor_usuario == "GERAL")
    
    if not acesso_total:
        flash("Acesso negado.")
        return redirect(url_for("index"))
    
    wb = load_workbook(ARQUIVO, data_only=True)
    alarmes = []
    if "Inspecoes" in wb.sheetnames:
        aba = wb["Inspecoes"]
        for row in aba.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 13: continue
            if str(row[6]).upper() == "ALARME" and str(row[12]).upper() == "PENDENTE":
                if acesso_total or str(row[5]).upper().strip() == setor_usuario:
                    alarmes.append({"id": row[0], "data": row[1], "eq": row[2], "item": row[3], "valor": row[4], "setor": row[5], "obs": row[7], "foto": row[8]})
    return render_template("tratativas.html", alarmes=alarmes, setor_nome=setor_usuario)

# ====================
# SALVAR TRATATIVAS
# ====================

@app.route("/salvar_tratativa", methods=["POST"])
@login_required
def salvar_tratativa():
    if session.get("nivel") == "ADMINFULL":
        flash("ADMINFULL não possui permissão para realizar tratativas.")
        return redirect(url_for("tratativas"))
        
    wb = load_workbook(ARQUIVO)
    aba = wb["Inspecoes"]
    eq_nome, item_nome = request.form.get("equipamento"), request.form.get("item")
    causa, acao = request.form.get("causa"), request.form.get("acao")
    
    for row in range(aba.max_row, 1, -1):
        if str(aba.cell(row=row, column=3).value) == eq_nome and str(aba.cell(row=row, column=4).value) == item_nome and str(aba.cell(row=row, column=13).value).upper() == "PENDENTE":
            aba.cell(row=row, column=11).value, aba.cell(row=row, column=12).value = causa, acao
            aba.cell(row=row, column=13).value = "CONCLUÍDO"
            break
    wb.save(ARQUIVO)
    flash("Tratativa registrada!")
    return redirect(url_for("tratativas"))

# ==============
# DASHBOARD
# ==============

@app.route("/dashboard")
@login_required
def dashboard():
    from datetime import datetime
    from openpyxl import load_workbook
    import os

    ARQUIVO = "base_dados.xlsx" 
    if not os.path.exists(ARQUIVO):
        return "Arquivo base_dados.xlsx não encontrado."

    wb = load_workbook(ARQUIVO, data_only=True, read_only=True)
    total = 0
    mapa_ultimos = {}
    agora = datetime.now()

    if "Inspecoes" in wb.sheetnames:
        aba = wb["Inspecoes"]

        for row in aba.iter_rows(min_row=2, values_only=True):
            if not row or row[2] is None:
                continue

            total += 1

            # --- CAPTURA ---
            data_raw        = row[1]
            equipamento     = str(row[2]).strip()
            item            = str(row[3]).strip()
            valor           = str(row[4]).strip() if row[4] is not None else "-"
            setor           = str(row[5]).strip()
            status_inspecao = str(row[6]).upper().strip() if row[6] else ""
            status_os       = str(row[16]).upper().strip() if len(row) > 16 and row[16] else "PENDENTE"
            
            # 🔥 CAPTURA COLUNA R (Índice 17)
            # Se estiver vazio, define como "NORMAL" para vermos algo na tela
            nivel_xlsx = str(row[17]).strip().upper() if len(row) > 17 and row[17] else "NORMAL"

            if status_inspecao == "ALARME" and status_os == "PENDENTE":
                if isinstance(data_raw, datetime):
                    data_obj = data_raw
                else:
                    try:
                        data_obj = datetime.strptime(str(data_raw), "%d/%m/%Y %H:%M")
                    except:
                        data_obj = agora

                chave = f"{equipamento}||{item}"

                if chave not in mapa_ultimos or data_obj > mapa_ultimos[chave]["_data_obj"]:
                    mapa_ultimos[chave] = {
                        "Data": data_obj.strftime("%d/%m/%Y %H:%M"),
                        "Equipamento": equipamento,
                        "Item": item,
                        "Valor": valor,
                        "Setor": setor,
                        "Nivel": nivel_xlsx, # Chave enviada ao HTML
                        "Dias": (agora - data_obj).days,
                        "_data_obj": data_obj 
                    }

    lista_final = list(mapa_ultimos.values())
    
    # Ordenação (Crítico e Alto no topo)
    pesos = {"CRÍTICO": 0, "ALTO": 1, "ALERTA": 1, "MÉDIO": 2, "BAIXO": 3, "NORMAL": 4}
    lista_final.sort(key=lambda x: pesos.get(x["Nivel"], 99))

    alarmes_por_setor = {}
    variaveis_alarme = {}
    for reg in lista_final:
        s, i = reg["Setor"], reg["Item"]
        alarmes_por_setor[s] = alarmes_por_setor.get(s, 0) + 1
        variaveis_alarme[i] = variaveis_alarme.get(i, 0) + 1

    return render_template(
        "dashboard_visual.html", 
        total=total, ok=total - len(lista_final), alarme=len(lista_final),
        lista_alarmes=lista_final,
        variaveis_alarme=variaveis_alarme,
        alarmes_por_setor=alarmes_por_setor,
        setor_nome=session.get("setor", "GERAL")
    )

# ==========
# SALVAR
# ==========

@app.route("/salvar", methods=["POST"])
@login_required
def salvar():
    if session.get("nivel") == "ADMINFULL":
        flash("Usuários ADMINFULL não têm permissão para salvar inspeções.")
        return redirect(url_for("index"))

    equipamento, setor, usuario = request.form.get("equipamento"), session.get("setor"), session.get("usuario")
    lat, lon = request.form.get("latitude"), request.form.get("longitude")
    data_agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    wb = load_workbook(ARQUIVO)
    aba_ins = wb["Inspecoes"]
    
    for campo, valor in request.form.items():
        if campo in ["equipamento", "latitude", "longitude"] or campo.startswith("status_real_") or campo.startswith("obs_"): continue
        status = request.form.get(f"status_real_{campo}", "OK")
        obs = request.form.get(f"obs_{campo}", "")
        nome_arquivo_foto = ""
        if f"foto_{campo}" in request.files:
            arquivo = request.files[f"foto_{campo}"]
            if arquivo and arquivo.filename != '':
                nome_arquivo_foto = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{campo}_{arquivo.filename}"
                arquivo.save(os.path.join(UPLOAD, nome_arquivo_foto))

        aba_ins.append([aba_ins.max_row, data_agora, equipamento, campo, valor, setor, status, obs, nome_arquivo_foto, usuario, "", "", "PENDENTE" if status == "ALARME" else "N/A", lat, lon])
        
    wb.save(ARQUIVO)
    flash(f"Inspeção de {equipamento} concluída!")
    return redirect(url_for("index"))

# ===========
# ROUTE
# ===========

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

verificar_arquivo_excel()
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)
